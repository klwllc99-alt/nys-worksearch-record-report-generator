from __future__ import annotations

import csv
import datetime as dt
import re
from collections import defaultdict
from dataclasses import dataclass
from io import BytesIO, StringIO
from pathlib import Path

from openpyxl import load_workbook


@dataclass(slots=True)
class WorkSearchEntry:
    date: dt.date
    position_or_activity: str
    business_name: str = ""
    person_contacted: str = ""
    method_of_contact: str = ""
    contact_information: str = ""
    result: str = ""
    record_type: str = "employer_contact"

    @property
    def activity(self) -> str:
        return self.position_or_activity

    @property
    def contact(self) -> str:
        return self.person_contacted

    @property
    def details(self) -> str:
        return self.result or self.contact_information


HEADER_ALIASES = {
    "date": {"date", "work_date", "activity_date", "work_search_date", "date_of_contact_activity", "date_of_contact", "date_of_activity"},
    "position_or_activity": {
        "position_applied_activity",
        "position_applied_for_activity_performed",
        "position_applied_for",
        "activity_performed",
        "position_or_activity",
        "activity",
        "work_search_activity",
        "method",
        "action",
        "type",
        "job_title",
        "position",
    },
    "business_name": {
        "business_employer_name",
        "business_website_agency",
        "business_name",
        "business",
        "company",
        "employer",
        "website",
        "agency",
    },
    "person_contacted": {
        "person_contacted_title",
        "person_contacted",
        "contact",
        "contact_name",
        "representative",
        "name_and_title_of_person_contacted",
    },
    "method_of_contact": {
        "method_of_contact",
        "contact_method",
        "how_contacted",
        "method",
    },
    "contact_information": {
        "contact_information",
        "contact_info",
        "address_telephone_email_website_url_fax_number",
        "phone_email_website",
    },
    "result": {
        "details_result",
        "details",
        "result",
        "notes",
        "outcome",
        "comments",
    },
    "record_type": {"record_type", "entry_type", "category", "section"},
}

REQUIRED_COLUMNS = ("date", "position_or_activity")


def normalize_header(value: object) -> str:
    text = str(value or "").strip().lower()
    text = re.sub(r"[^a-z0-9]+", "_", text)
    return text.strip("_")


def _coerce_text(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, str):
        return value.strip()
    return str(value).strip()


def _parse_date(value: object) -> dt.date:
    if isinstance(value, dt.datetime):
        return value.date()
    if isinstance(value, dt.date):
        return value

    text = _coerce_text(value)
    if not text:
        raise ValueError("Missing date value in one or more rows.")

    for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%m/%d/%y", "%b %d %Y", "%B %d %Y"):
        try:
            return dt.datetime.strptime(text, fmt).date()
        except ValueError:
            continue

    raise ValueError(f"Unsupported date format: {text}")


def _week_ending_sunday(value: dt.date) -> dt.date:
    return value + dt.timedelta(days=(6 - value.weekday()) % 7)


def _resolve_header_indexes(headers: list[object]) -> dict[str, int]:
    indexes: dict[str, int] = {}
    for idx, header in enumerate(headers):
        normalized = normalize_header(header)
        for canonical, aliases in HEADER_ALIASES.items():
            if normalized == canonical or normalized in aliases:
                indexes.setdefault(canonical, idx)

    missing = [column for column in REQUIRED_COLUMNS if column not in indexes]
    if missing:
        readable = ", ".join(missing)
        raise ValueError(f"Missing required columns: {readable}")

    return indexes


def _infer_record_type(raw_value: str, business_name: str, person_contacted: str, contact_information: str, result: str) -> str:
    normalized = raw_value.strip().lower().replace("-", "_").replace(" ", "_")
    if normalized in {"other", "other_activity", "other_activities", "activity"}:
        return "other_activity"
    if normalized in {"employer", "employer_contact", "business", "business_contact", "job_contact"}:
        return "employer_contact"
    if business_name or person_contacted or contact_information or result:
        return "employer_contact"
    return "other_activity"


def _build_entry(values: dict[str, object]) -> WorkSearchEntry:
    business_name = _coerce_text(values.get("business_name"))
    person_contacted = _coerce_text(values.get("person_contacted"))
    contact_information = _coerce_text(values.get("contact_information"))
    result = _coerce_text(values.get("result"))

    return WorkSearchEntry(
        date=_parse_date(values.get("date")),
        position_or_activity=_coerce_text(values.get("position_or_activity")),
        business_name=business_name,
        person_contacted=person_contacted,
        method_of_contact=_coerce_text(values.get("method_of_contact")),
        contact_information=contact_information,
        result=result,
        record_type=_infer_record_type(
            _coerce_text(values.get("record_type")),
            business_name,
            person_contacted,
            contact_information,
            result,
        ),
    )


def _parse_csv(payload: bytes) -> list[WorkSearchEntry]:
    text = payload.decode("utf-8-sig")
    reader = csv.DictReader(StringIO(text))
    if not reader.fieldnames:
        raise ValueError("CSV file is missing a header row.")

    header_lookup: dict[str, str] = {}
    for header in reader.fieldnames:
        normalized = normalize_header(header)
        for canonical, aliases in HEADER_ALIASES.items():
            if normalized == canonical or normalized in aliases:
                header_lookup.setdefault(canonical, header)

    missing = [column for column in REQUIRED_COLUMNS if column not in header_lookup]
    if missing:
        readable = ", ".join(missing)
        raise ValueError(f"Missing required columns: {readable}")

    entries: list[WorkSearchEntry] = []
    for row in reader:
        if not row or not any(_coerce_text(value) for value in row.values()):
            continue

        values = {
            canonical: row.get(source_header)
            for canonical, source_header in header_lookup.items()
        }
        entries.append(_build_entry(values))

    return entries


def _parse_xlsx(payload: bytes) -> list[WorkSearchEntry]:
    workbook = load_workbook(filename=BytesIO(payload), read_only=True, data_only=True)
    sheet = workbook.active
    row_iter = sheet.iter_rows(values_only=True)

    try:
        header_row = list(next(row_iter))
    except StopIteration as exc:
        raise ValueError("XLSX file is empty.") from exc

    header_indexes = _resolve_header_indexes(header_row)
    entries: list[WorkSearchEntry] = []

    for row in row_iter:
        if not row or not any(_coerce_text(value) for value in row):
            continue

        values = {
            canonical: row[index] if index < len(row) else None
            for canonical, index in header_indexes.items()
        }
        entries.append(_build_entry(values))

    return entries


def parse_work_search_file(filename: str, payload: bytes) -> dict[dt.date, list[WorkSearchEntry]]:
    extension = Path(filename).suffix.lower()
    if extension == ".csv":
        entries = _parse_csv(payload)
    elif extension == ".xlsx":
        entries = _parse_xlsx(payload)
    else:
        raise ValueError("Only .csv and .xlsx files are supported.")

    if not entries:
        raise ValueError("No valid work search rows were found in the upload.")

    grouped: dict[dt.date, list[WorkSearchEntry]] = defaultdict(list)
    for entry in sorted(entries, key=lambda item: (item.date, item.business_name.lower(), item.position_or_activity.lower())):
        grouped[_week_ending_sunday(entry.date)].append(entry)

    return dict(sorted(grouped.items(), key=lambda item: item[0]))
